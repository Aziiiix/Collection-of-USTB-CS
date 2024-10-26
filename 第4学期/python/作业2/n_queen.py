import chessboard
import queen
import openpyxl

n=8

ans1= [None for i in range(n)]
ans2= [None for i in range(n)]
ans1,ans2=queen.fun_queen(n,ans1,ans2)

chessboard.fun_chessboard(n)

wb=openpyxl.load_workbook('chessboard.xlsx')
sheet=wb.active
for i in range(0,n):
    sheet.cell(i+1,ans1[i]+1).value='\u2655'
for i in range(0,n):
    sheet.cell(n+2+i, ans2[i] + 1).value = '\u2655'
wb.save('n_queen.xlsx')