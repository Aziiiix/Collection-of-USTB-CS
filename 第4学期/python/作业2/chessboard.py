import openpyxl
from openpyxl.styles import PatternFill,Font,colors,Alignment

def fun_chessboard(n):
    wb = openpyxl.Workbook()
    sheet = wb.active

    for i in range(1, 2*n+3):
        row = sheet.row_dimensions[i]
        row.height = 35

    for i in range(1, n+1):
        col = sheet.column_dimensions[chr(64 + i)]
        col.width = 10

    blackfill = PatternFill("solid", fgColor="000000")
    purplefill = PatternFill("solid", fgColor="8B658B")
    writefont=Font(name='宋体',size=28,color=colors.BLACK)
    blackfont = Font(name='宋体', size=28, color=colors.WHITE)

    for i in range(1, n+1):
        if (i % 2 == 0):
            for j in range(1, n+1):
                sheet.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
                if (j % 2 == 0):
                    sheet.cell(i,j).font=writefont
                    continue
                else:
                    sheet.cell(i, j).fill = blackfill
                    sheet.cell(i,j).font=blackfont
        else:
            for j in range(1, n+1):
                sheet.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
                if (j % 2 == 0):
                    sheet.cell(i, j).fill = blackfill
                    sheet.cell(i, j).font = blackfont
                else:
                    sheet.cell(i, j).font = writefont
                    continue

    for i in range(1,n+1):
        sheet.cell(n+1,i).fill=purplefill

    for i in range(n+2,2*n+2):
        if (i % 2 == 0):
            for j in range(1, n + 1):
                sheet.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
                if (j % 2 == 0):
                    sheet.cell(i, j).fill = blackfill
                    sheet.cell(i, j).font = blackfont
                else:
                    sheet.cell(i, j).font = writefont
                    continue
        else:
            for j in range(1, n + 1):
                sheet.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
                if (j % 2 == 0):
                    sheet.cell(i, j).font = writefont
                    continue
                else:
                    sheet.cell(i, j).fill = blackfill
                    sheet.cell(i, j).font = blackfont

    for i in range(1,n+1):
        sheet.cell(2*n+2,i).fill=purplefill

    for i in range(1,2*n+3):
        sheet.cell(i,n+1).fill=purplefill

    wb.save('chessboard.xlsx')
